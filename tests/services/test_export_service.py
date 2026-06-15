"""Unit tests for ExportService — backend-agnostic row building."""

from __future__ import annotations

import csv
import io
from datetime import datetime

import openpyxl

from rentivo.export.serializers import rows_to_csv_bytes, rows_to_xlsx_bytes
from rentivo.models.bill import Bill, BillStatus
from rentivo.models.billing import Billing
from rentivo.services.export_service import ExportService


def _billing() -> Billing:
    return Billing(id=1, uuid="bil-u", name="Apt 101")


def _bill(**overrides) -> Bill:
    defaults = dict(
        id=1,
        uuid="bill-u",
        billing_id=1,
        reference_month="2025-03",
        total_amount=285050,
        due_date="10/04/2025",
        status=BillStatus.PAID.value,
        status_updated_at=datetime(2025, 4, 9, 14, 30),
    )
    defaults.update(overrides)
    return Bill(**defaults)


class TestBuildRows:
    def test_headers_are_pt_br(self):
        assert ExportService().HEADERS == [
            "Mês de referência",
            "Cobrança",
            "Valor (R$)",
            "Valor formatado",
            "Status",
            "Vencimento",
            "Pago/Atualizado em",
        ]

    def test_row_values(self):
        rows = ExportService().build_rows(_billing(), [_bill()])
        assert rows == [
            [
                "Março/2025",
                "Apt 101",
                2850.50,
                "R$ 2.850,50",
                "Pago",
                "10/04/2025",
                "09/04/2025 14:30",
            ]
        ]

    def test_amount_column_is_numeric_float(self):
        rows = ExportService().build_rows(_billing(), [_bill()])
        assert isinstance(rows[0][2], float)

    def test_unknown_status_falls_back_to_raw_value(self):
        rows = ExportService().build_rows(_billing(), [_bill(status="weird")])
        assert rows[0][4] == "weird"

    def test_missing_optional_fields_render_empty(self):
        rows = ExportService().build_rows(_billing(), [_bill(due_date=None, status_updated_at=None)])
        assert rows[0][5] == ""
        assert rows[0][6] == ""

    def test_empty_bill_list_yields_no_rows(self):
        assert ExportService().build_rows(_billing(), []) == []


HEADERS = ["A", "B"]
ROWS = [["x", 1.5], ["y", 2.0]]


class TestCsvSerializer:
    def test_csv_round_trips_header_and_rows(self):
        data = rows_to_csv_bytes(HEADERS, ROWS)
        assert isinstance(data, bytes)
        text = data.decode("utf-8-sig")
        parsed = list(csv.reader(io.StringIO(text)))
        assert parsed[0] == ["A", "B"]
        assert parsed[1] == ["x", "1.5"]
        assert parsed[2] == ["y", "2.0"]

    def test_csv_starts_with_utf8_bom(self):
        data = rows_to_csv_bytes(HEADERS, ROWS)
        assert data.startswith(b"\xef\xbb\xbf")


class TestXlsxSerializer:
    def test_xlsx_is_a_valid_workbook_with_header_and_rows(self):
        data = rows_to_xlsx_bytes(HEADERS, ROWS)
        assert isinstance(data, bytes)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        assert [c.value for c in ws[1]] == ["A", "B"]
        assert [c.value for c in ws[2]] == ["x", 1.5]
        assert ws[2][1].value == 2.0 or ws[3][1].value == 2.0

    def test_xlsx_preserves_numeric_cells(self):
        data = rows_to_xlsx_bytes(HEADERS, ROWS)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        assert isinstance(ws["B2"].value, float)


class TestFormulaInjectionNeutralization:
    def test_csv_neutralizes_equals_formula(self):
        data = rows_to_csv_bytes(["A"], [["=cmd|calc!A1"]])
        parsed = list(csv.reader(io.StringIO(data.decode("utf-8-sig"))))
        assert parsed[1][0] == "'=cmd|calc!A1"

    def test_csv_neutralizes_plus_minus_at(self):
        data = rows_to_csv_bytes(["A"], [["+1"], ["-1"], ["@x"]])
        parsed = list(csv.reader(io.StringIO(data.decode("utf-8-sig"))))
        assert parsed[1][0] == "'+1"
        assert parsed[2][0] == "'-1"
        assert parsed[3][0] == "'@x"

    def test_csv_leaves_normal_string_unchanged(self):
        data = rows_to_csv_bytes(["A"], [["Apartamento 101"]])
        parsed = list(csv.reader(io.StringIO(data.decode("utf-8-sig"))))
        assert parsed[1][0] == "Apartamento 101"

    def test_xlsx_neutralizes_equals_formula(self):
        data = rows_to_xlsx_bytes(["A"], [["=cmd|calc!A1"]])
        ws = openpyxl.load_workbook(io.BytesIO(data)).active
        assert ws["A2"].value == "'=cmd|calc!A1"

    def test_xlsx_neutralizes_plus_minus_at(self):
        data = rows_to_xlsx_bytes(["A"], [["+1"], ["-1"], ["@x"]])
        ws = openpyxl.load_workbook(io.BytesIO(data)).active
        assert ws["A2"].value == "'+1"
        assert ws["A3"].value == "'-1"
        assert ws["A4"].value == "'@x"

    def test_xlsx_leaves_normal_string_unchanged(self):
        data = rows_to_xlsx_bytes(["A"], [["Apartamento 101"]])
        ws = openpyxl.load_workbook(io.BytesIO(data)).active
        assert ws["A2"].value == "Apartamento 101"

    def test_xlsx_leaves_numeric_cell_as_number(self):
        data = rows_to_xlsx_bytes(["A", "B"], [["=x", 2850.50]])
        ws = openpyxl.load_workbook(io.BytesIO(data)).active
        assert isinstance(ws["B2"].value, float)
        assert ws["A2"].value == "'=x"
