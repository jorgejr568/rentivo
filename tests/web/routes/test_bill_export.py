"""Integration tests for the bill export route (CSV + XLSX)."""

from __future__ import annotations

import csv
import io
from unittest.mock import patch

import openpyxl

from rentivo.models.audit_log import AuditEventType
from rentivo.storage.local import LocalStorage
from tests.web.conftest import create_billing_in_db, generate_bill_in_db, get_audit_logs


class TestBillExport:
    def test_export_csv_default_format(self, auth_client, test_engine, tmp_path):
        billing = create_billing_in_db(test_engine)
        with patch("web.deps.get_storage", return_value=LocalStorage(str(tmp_path))):
            generate_bill_in_db(test_engine, billing, tmp_path)
            response = auth_client.get(f"/billings/{billing.uuid}/bills/export")
        assert response.status_code == 200
        assert response.headers["content-type"].startswith("text/csv")
        assert "attachment" in response.headers["content-disposition"]
        assert ".csv" in response.headers["content-disposition"]
        text = response.content.decode("utf-8-sig")
        parsed = list(csv.reader(io.StringIO(text)))
        assert parsed[0][0] == "Mês de referência"
        assert parsed[1][1] == "Apt 101"

    def test_export_csv_explicit_format(self, auth_client, test_engine, tmp_path):
        billing = create_billing_in_db(test_engine)
        with patch("web.deps.get_storage", return_value=LocalStorage(str(tmp_path))):
            generate_bill_in_db(test_engine, billing, tmp_path)
            response = auth_client.get(f"/billings/{billing.uuid}/bills/export?format=csv")
        assert response.status_code == 200
        assert response.headers["content-type"].startswith("text/csv")

    def test_export_xlsx_format(self, auth_client, test_engine, tmp_path):
        billing = create_billing_in_db(test_engine)
        with patch("web.deps.get_storage", return_value=LocalStorage(str(tmp_path))):
            generate_bill_in_db(test_engine, billing, tmp_path)
            response = auth_client.get(f"/billings/{billing.uuid}/bills/export?format=xlsx")
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert ".xlsx" in response.headers["content-disposition"]
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        assert ws[1][0].value == "Mês de referência"
        assert ws[2][1].value == "Apt 101"

    def test_export_unknown_format_falls_back_to_csv(self, auth_client, test_engine, tmp_path):
        billing = create_billing_in_db(test_engine)
        with patch("web.deps.get_storage", return_value=LocalStorage(str(tmp_path))):
            generate_bill_in_db(test_engine, billing, tmp_path)
            response = auth_client.get(f"/billings/{billing.uuid}/bills/export?format=pdf")
        assert response.status_code == 200
        assert response.headers["content-type"].startswith("text/csv")

    def test_export_empty_billing_returns_header_only_csv(self, auth_client, test_engine):
        billing = create_billing_in_db(test_engine)
        response = auth_client.get(f"/billings/{billing.uuid}/bills/export")
        assert response.status_code == 200
        parsed = list(csv.reader(io.StringIO(response.content.decode("utf-8-sig"))))
        assert parsed[0][0] == "Mês de referência"
        assert len(parsed) == 1

    def test_export_records_audit_event(self, auth_client, test_engine, tmp_path):
        billing = create_billing_in_db(test_engine)
        with patch("web.deps.get_storage", return_value=LocalStorage(str(tmp_path))):
            generate_bill_in_db(test_engine, billing, tmp_path)
            auth_client.get(f"/billings/{billing.uuid}/bills/export?format=xlsx")
        logs = get_audit_logs(test_engine, event_type=AuditEventType.BILLING_EXPORT)
        assert len(logs) == 1
        assert logs[0].new_state == {"format": "xlsx", "bill_count": 1}

    def test_export_billing_not_found(self, auth_client):
        response = auth_client.get("/billings/nonexistent/bills/export", follow_redirects=False)
        assert response.status_code == 302

    def test_export_other_users_billing_denied(self, auth_client, test_engine):
        from rentivo.encryption.base64 import Base64Backend
        from rentivo.models.user import User
        from rentivo.repositories.sqlalchemy import SQLAlchemyUserRepository

        with test_engine.connect() as conn:
            other = SQLAlchemyUserRepository(conn, Base64Backend()).create(
                User(email="exp_other@example.com", password_hash="h")
            )
        billing = create_billing_in_db(test_engine, owner_type="user", owner_id=other.id)
        response = auth_client.get(f"/billings/{billing.uuid}/bills/export", follow_redirects=False)
        assert response.status_code == 302


class TestExportSlug:
    def test_preserves_accents_sao_joao(self):
        from web.routes.bill import _export_slug

        assert _export_slug("São João") == "sao-joao"

    def test_preserves_accents_atica(self):
        from web.routes.bill import _export_slug

        assert _export_slug("Ática") == "atica"

    def test_empty_falls_back_to_cobranca(self):
        from web.routes.bill import _export_slug

        assert _export_slug("!!!") == "cobranca"
